"""Proposal-5 trial scheduling (v2 — fixed xlsx).

试次素材库 (12 grid : 12 loop : 6 tie = 30 pairs)。
每 session 20 试次，配额 grid:loop:tie = 8:8:4。
5 个 session 的固定试次序列预先生成并保存到 trial_schedule.xlsx，
运行时直接读取，不再使用 random seed。
"""

from __future__ import annotations

import json
import random
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Literal, Optional, Sequence, Tuple

Category = Literal["grid", "loop", "tie"]


@dataclass(frozen=True)
class PairSpec:
    pair_id: str
    category: Category
    start: int
    goal: int
    d_grid: int
    d_loop: int
    multiplicity: int
    grid_paths: int
    loop_paths: int


TRIALS_PER_SESSION = 20
DEFAULT_BLOCK_QUOTAS: Dict[Category, int] = {"grid": 8, "loop": 8, "tie": 4}

# ── 素材库 12 grid + 12 loop + 6 tie = 30 pairs ────────────────
_PAIR_ROWS: Dict[Category, List[Tuple[str, int, int, int, int, int]]] = {
    # Grid pairs (d_grid < d_loop): 12 pairs
    "grid": [
        ("G01", 1, 5, 2, 3, 2),
        ("G02", 2, 6, 2, 3, 2),
        ("G03", 3, 5, 2, 3, 2),
        ("G04", 4, 2, 2, 3, 2),
        ("G05", 5, 1, 2, 3, 2),
        ("G06", 5, 3, 2, 3, 2),
        ("G07", 5, 7, 2, 3, 2),
        ("G08", 5, 9, 2, 3, 2),
        ("G09", 6, 8, 2, 3, 2),
        ("G10", 7, 5, 2, 3, 2),
        ("G11", 9, 3, 2, 3, 1),
        ("G12", 9, 5, 2, 3, 2),
    ],
    # Loop pairs (d_loop < d_grid): 12 pairs
    "loop": [
        ("L01", 1, 6, 3, 2, 1),
        ("L02", 1, 9, 4, 2, 1),
        ("L03", 2, 9, 3, 2, 1),
        ("L04", 3, 7, 4, 2, 1),
        ("L05", 3, 8, 3, 2, 1),
        ("L06", 4, 3, 3, 2, 1),
        ("L07", 6, 7, 3, 2, 1),
        ("L08", 7, 2, 3, 2, 1),
        ("L09", 7, 3, 4, 2, 1),
        ("L10", 8, 1, 3, 2, 1),
        ("L11", 9, 1, 4, 2, 1),
        ("L12", 9, 4, 3, 2, 1),
    ],
    # Tie pairs (d_grid == d_loop): 6 pairs
    "tie": [
        ("T01", 1, 8, 3, 3, 4),
        ("T02", 2, 7, 3, 3, 4),
        ("T03", 3, 4, 3, 3, 4),
        ("T04", 4, 9, 3, 3, 4),
        ("T05", 6, 1, 3, 3, 4),
        ("T06", 7, 6, 3, 3, 4),
    ],
}

_PAIR_CATALOG: Optional[Dict[Category, List[PairSpec]]] = None


def _row_to_spec(category: Category, row: Tuple[str, int, int, int, int, int]) -> PairSpec:
    pair_id, start, goal, d_grid, d_loop, multiplicity = row
    if category == "grid":
        grid_paths, loop_paths = multiplicity, 0
    elif category == "loop":
        grid_paths, loop_paths = 0, multiplicity
    else:
        grid_paths, loop_paths = multiplicity, multiplicity
    return PairSpec(
        pair_id=pair_id,
        category=category,
        start=start,
        goal=goal,
        d_grid=d_grid,
        d_loop=d_loop,
        multiplicity=multiplicity,
        grid_paths=grid_paths,
        loop_paths=loop_paths,
    )


def get_pair_catalog() -> Dict[Category, List[PairSpec]]:
    global _PAIR_CATALOG
    if _PAIR_CATALOG is None:
        _PAIR_CATALOG = {
            category: [_row_to_spec(category, row) for row in rows]
            for category, rows in _PAIR_ROWS.items()
        }
    return _PAIR_CATALOG


def _weighted_choice(
    pairs: Sequence[PairSpec],
    rng: random.Random,
    last_pair_id: Optional[str],
) -> PairSpec:
    options = [pair for pair in pairs if pair.pair_id != last_pair_id]
    if not options:
        options = list(pairs)
    weights = [pair.multiplicity for pair in options]
    return rng.choices(options, weights=weights, k=1)[0]


def _build_category_sequence(quotas: Dict[Category, int], rng: random.Random) -> List[Category]:
    remaining = dict(quotas)
    sequence: List[Category] = []
    for _ in range(sum(remaining.values())):
        candidates = [category for category, left in remaining.items() if left > 0]
        weights = [remaining[category] for category in candidates]
        chosen = rng.choices(candidates, weights=weights, k=1)[0]
        sequence.append(chosen)
        remaining[chosen] -= 1
    return sequence


def generate_navigation_trials(
    seed: int,
    trial_count: int = TRIALS_PER_SESSION,
    quotas: Optional[Dict[Category, int]] = None,
) -> List[Dict[str, object]]:
    quotas = quotas or DEFAULT_BLOCK_QUOTAS
    if sum(quotas.values()) != trial_count:
        raise ValueError("Quota sum must equal trial_count")

    catalog = get_pair_catalog()
    rng = random.Random(seed)
    category_sequence = _build_category_sequence(quotas, rng)

    trials: List[Dict[str, object]] = []
    last_pair_id: Optional[str] = None
    for block_index, category in enumerate(category_sequence, start=1):
        chosen = _weighted_choice(catalog[category], rng, last_pair_id)
        last_pair_id = chosen.pair_id
        trials.append(
            {
                "index": block_index,
                "pair_id": chosen.pair_id,
                "category": chosen.category,
                "start": chosen.start,
                "goal": chosen.goal,
                "d_grid": chosen.d_grid,
                "d_loop": chosen.d_loop,
                "grid_paths": chosen.grid_paths,
                "loop_paths": chosen.loop_paths,
                "multiplicity": chosen.multiplicity,
            }
        )
    return trials


CRAFTING_MOCK_RECIPES = [
    "smelt_ingot",
    "craft_table",
    "brew_potion",
    "assemble_cart",
    "cook_stew",
    "forge_sword",
]


def _make_crafting_trial(
    pair_trial: Dict[str, object],
    rng: random.Random,
    session_number: int,
    craft_index: int,
) -> Dict[str, object]:
    return {
        "index": craft_index,
        "block_index": int(pair_trial["index"]),
        "task_id": f"S{session_number}_craft_{craft_index:02d}",
        "recipe": rng.choice(CRAFTING_MOCK_RECIPES),
        "pair_id": pair_trial["pair_id"],
        "category": pair_trial["category"],
        "start": pair_trial["start"],
        "goal": pair_trial["goal"],
        "d_grid": pair_trial["d_grid"],
        "d_loop": pair_trial["d_loop"],
        "grid_paths": pair_trial["grid_paths"],
        "loop_paths": pair_trial["loop_paths"],
        "multiplicity": pair_trial["multiplicity"],
    }


# ── Session 结构：无论 order 如何，固定使用 xlsx 中的试次 ──────
_SESSION_DOMAINS = [
    {"session": 1, "domain_nav_first": "navigation", "domain_craft_first": "crafting"},
    {"session": 2, "domain_nav_first": "crafting",   "domain_craft_first": "navigation"},
    {"session": 3, "domain_nav_first": "navigation", "domain_craft_first": "crafting"},
    {"session": 4, "domain_nav_first": "crafting",   "domain_craft_first": "navigation"},
    {"session": 5, "domain_nav_first": "mixed",      "domain_craft_first": "mixed"},
]


def _xlsx_path() -> Path:
    return Path(__file__).resolve().parents[2] / "trial_schedule.xlsx"


def load_schedule_from_xlsx(xlsx_path: Path | str | None = None) -> List[List[Dict[str, object]]]:
    """从 xlsx 加载 5 session 的固定试次序列。

    返回 ``sessions[0..4]``，每个元素是长度 25 的 trial list。
    """
    import openpyxl

    xlsx_path = Path(xlsx_path) if xlsx_path else _xlsx_path()
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    sessions: List[List[Dict[str, object]]] = []

    for sn in range(1, 6):
        ws = wb[f"Session{sn}"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        trials: List[Dict[str, object]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                break
            trial = dict(zip(headers, row))
            trials.append(trial)
        sessions.append(trials)

    wb.close()
    return sessions


def build_session_schedule(order: str) -> Dict[str, object]:
    """从 xlsx 构建运行时 schedule，兼容原有调度器接口。"""
    xlsx = _xlsx_path()
    if not xlsx.is_file():
        raise FileNotFoundError(
            f"固定试次表不存在: {xlsx}\n请先运行 generate_trial_xlsx() 生成。"
        )
    all_sessions = load_schedule_from_xlsx(xlsx)
    crafting_rng = random.Random(42)
    sessions: List[Dict[str, object]] = []

    for idx, spec in enumerate(_SESSION_DOMAINS):
        sn = spec["session"]
        domain_key = "domain_nav_first" if order == "navigation-first" else "domain_craft_first"
        domain = spec[domain_key]
        raw_trials = all_sessions[idx]

        navigation_trials: List[Dict[str, object]] = []
        crafting_trials: List[Dict[str, object]] = []
        combined_order: List[Dict[str, object]] = []

        if domain == "navigation":
            for i, t in enumerate(raw_trials, 1):
                trial = _normalize_trial(t, i)
                navigation_trials.append(trial)
                combined_order.append({"type": "navigation", "trial": trial})

        elif domain == "crafting":
            for i, t in enumerate(raw_trials, 1):
                pair_trial = _normalize_trial(t, i)
                craft_trial = _make_crafting_trial(pair_trial, crafting_rng, sn, i)
                crafting_trials.append(craft_trial)
                combined_order.append({"type": "crafting", "trial": craft_trial})

        elif domain == "mixed":
            nav_idx = 0
            craft_idx = 0
            for i, t in enumerate(raw_trials, 1):
                pair_trial = _normalize_trial(t, i)
                # 混合 session：奇数试次 = crafting，偶数试次 = navigation
                # （crafting 先行交替）
                if i % 2 == 1:
                    craft_idx += 1
                    craft_trial = _make_crafting_trial(pair_trial, crafting_rng, sn, craft_idx)
                    crafting_trials.append(craft_trial)
                    combined_order.append({"type": "crafting", "trial": craft_trial})
                else:
                    nav_idx += 1
                    nav_trial = {**pair_trial, "index": nav_idx}
                    navigation_trials.append(nav_trial)
                    combined_order.append({"type": "navigation", "trial": nav_trial})

        sessions.append({
            "session": sn,
            "domain": domain,
            "seed": 0,
            "navigation_trials": navigation_trials,
            "crafting_trials": crafting_trials,
            "combined_order": combined_order,
        })

    catalog = get_pair_catalog()
    return {
        "order": order,
        "sessions": sessions,
        "catalog_sizes": {category: len(pairs) for category, pairs in catalog.items()},
        "default_block_quotas": DEFAULT_BLOCK_QUOTAS,
    }


def _normalize_trial(raw: Dict[str, object], index: int) -> Dict[str, object]:
    """xlsx 原始行 → 标准 trial dict。"""
    return {
        "index": index,
        "pair_id": str(raw.get("pair_id", "")),
        "category": str(raw.get("category", "")),
        "start": int(raw.get("start", 0)),
        "goal": int(raw.get("goal", 0)),
        "d_grid": int(raw.get("d_grid", 0)),
        "d_loop": int(raw.get("d_loop", 0)),
        "grid_paths": int(raw.get("grid_paths", 0)),
        "loop_paths": int(raw.get("loop_paths", 0)),
        "multiplicity": int(raw.get("multiplicity", 0)),
        "block_index": index,
    }


# ── xlsx 生成 ─────────────────────────────────────────────
def generate_trial_xlsx(
    output_path: Path | str | None = None,
    seed: int = 20260408,
) -> Path:
    """生成 5 session × 25 试次的固定 xlsx 文件。"""
    import openpyxl

    output_path = Path(output_path) if output_path else _xlsx_path()
    wb = openpyxl.Workbook()
    # 删除默认 sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    headers = [
        "index", "pair_id", "category", "start", "goal",
        "d_grid", "d_loop", "grid_paths", "loop_paths", "multiplicity",
    ]

    for sn in range(1, 6):
        session_seed = seed + sn
        trials = generate_navigation_trials(
            seed=session_seed,
            trial_count=TRIALS_PER_SESSION,
            quotas=DEFAULT_BLOCK_QUOTAS,
        )
        ws = wb.create_sheet(title=f"Session{sn}")
        ws.append(headers)
        for trial in trials:
            ws.append([trial[h] for h in headers])

    wb.save(str(output_path))
    return output_path


def save_schedule(schedule: Dict[str, object], output_path: Path | str) -> None:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as fh:
        json.dump(schedule, fh, ensure_ascii=False, indent=2)
