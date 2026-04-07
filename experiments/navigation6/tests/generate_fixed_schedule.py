#!/usr/bin/env python3
"""Generate deterministic trial schedules and export to xlsx.

This script replaces the previous seed-based on-the-fly sampling approach.
It produces a **fixed** pseudo-random trial sequence for all 5 sessions
(120 trials total: 60 navigation + 60 crafting) and persists the result
as xlsx files under ``fixed_schedules/``.

Pair generation rules
---------------------
* The pair catalog is copied verbatim from ``trial_schedule.py`` (Tables 3-5):
  24 grid-dominant, 12 loop-dominant, 8 tie pairs — 44 unique pairs in total.
* Additional *instances* (repeats of existing catalog pairs) are selected
  round-robin within each category so that every unique pair is used at least
  once before any pair repeats.
* Per-session quotas follow ratio **7 : 7 : 10** (grid : loop : tie).
  - 24-trial sessions: 7 grid + 7 loop + 10 tie = 24
  - 12-trial session:  4 grid + 3 loop + 5 tie  = 12  (closest integer split)
* Navigation and crafting each get their own independent 60-trial sequence.
* Trial order within each session is pseudo-randomly shuffled with a fixed
  seed so the order is reproducible but not simply category-sorted.

Session layout (identical for navigation-first and crafting-first;
the *domain assignment* differs)
---------------------------------
Navigation domain: Session A (24) + Session B (24) + Session C-half (12) = 60
Crafting domain:   Session D (24) + Session E (24) + Session F-half (12) = 60

For **navigation-first**: S1=nav(24), S2=craft(24), S3=nav(24), S4=craft(24),
S5=mixed(12 nav + 12 craft alternating, nav first).

For **crafting-first**: S1=craft(24), S2=nav(24), S3=craft(24), S4=nav(24),
S5=mixed(12 craft + 12 nav alternating, craft first).
"""

from __future__ import annotations

import random
import sys
from pathlib import Path
from typing import Dict, List, Literal, Sequence, Tuple

_THIS_FILE = Path(__file__).resolve()
_PROJECT_ROOT = _THIS_FILE.parents[3]
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

from experiments.navigation6.tests.trial_schedule import (
    Category,
    PairSpec,
    get_pair_catalog,
    CRAFTING_MOCK_RECIPES,
)

# ── constants ─────────────────────────────────────────────────────────────

# Fixed seeds for reproducibility
_NAV_SEED = 20260407
_CRAFT_SEED = 20260408
_SHUFFLE_SEED_BASE = 99900

# Per-session quotas: grid : loop : tie
_QUOTA_24: Dict[Category, int] = {"grid": 7, "loop": 7, "tie": 10}
_QUOTA_12: Dict[Category, int] = {"grid": 4, "loop": 3, "tie": 5}

# Session sizes for each domain (navigation or crafting): 24 + 24 + 12 = 60
_SESSION_SIZES: List[int] = [24, 24, 12]

_FIXED_SCHEDULES_DIR = _THIS_FILE.parent / "fixed_schedules"


# ── pair selection ────────────────────────────────────────────────────────

def _select_pairs_round_robin(
    catalog_pairs: List[PairSpec],
    count: int,
    rng: random.Random,
) -> List[PairSpec]:
    """Select *count* pair instances from *catalog_pairs* via round-robin.

    Every unique pair is used once before any pair repeats.  Within each
    round the order is shuffled so that the same pair does not always
    appear at the same position.
    """
    pool = list(catalog_pairs)
    selected: List[PairSpec] = []
    while len(selected) < count:
        batch = list(pool)
        rng.shuffle(batch)
        needed = count - len(selected)
        selected.extend(batch[:needed])
    return selected


def _generate_domain_trials(
    seed: int,
    domain_label: str,
) -> List[Dict[str, object]]:
    """Return 60 trial dicts for one domain (navigation or crafting).

    Trials are grouped into three blocks of sizes 24, 24, 12.  Each block
    respects the 7:7:10 category quota (scaled for the 12-block).
    """
    catalog = get_pair_catalog()
    rng = random.Random(seed)

    all_trials: List[Dict[str, object]] = []
    global_index = 0

    for block_idx, size in enumerate(_SESSION_SIZES):
        quota = _QUOTA_24 if size == 24 else _QUOTA_12
        block_trials: List[Dict[str, object]] = []

        for category in ("grid", "loop", "tie"):
            n = quota[category]
            pairs = _select_pairs_round_robin(catalog[category], n, rng)
            for pair in pairs:
                global_index += 1
                trial: Dict[str, object] = {
                    "global_index": global_index,
                    "block": block_idx + 1,
                    "domain": domain_label,
                    "pair_id": pair.pair_id,
                    "category": pair.category,
                    "start": pair.start,
                    "goal": pair.goal,
                    "d_grid": pair.d_grid,
                    "d_loop": pair.d_loop,
                    "grid_paths": pair.grid_paths,
                    "loop_paths": pair.loop_paths,
                    "multiplicity": pair.multiplicity,
                }
                block_trials.append(trial)

        # Pseudo-random shuffle within block
        shuffle_rng = random.Random(_SHUFFLE_SEED_BASE + seed + block_idx)
        shuffle_rng.shuffle(block_trials)

        # Re-number block_index after shuffle
        for i, trial in enumerate(block_trials, start=1):
            trial["block_index"] = i

        all_trials.extend(block_trials)

    return all_trials


# ── crafting trial decoration ─────────────────────────────────────────────

def _decorate_crafting_trials(
    trials: List[Dict[str, object]],
    seed: int,
) -> None:
    """Add crafting-specific fields to each trial dict (in-place)."""
    rng = random.Random(seed + 200_000)
    for i, trial in enumerate(trials, start=1):
        trial["task_id"] = f"craft_{i:03d}"
        trial["recipe"] = rng.choice(CRAFTING_MOCK_RECIPES)


# ── five-session assembly ─────────────────────────────────────────────────

def _build_five_sessions(
    order: Literal["navigation-first", "crafting-first"],
    nav_trials: List[Dict[str, object]],
    craft_trials: List[Dict[str, object]],
) -> List[Dict[str, object]]:
    """Assemble 120 trials into 5 sessions following the given order.

    Returns a flat list of trial dicts, each annotated with ``session``,
    ``session_domain``, and ``session_index`` fields.
    """
    nav_first = order == "navigation-first"

    # Split each domain's 60 trials into blocks: [0:24], [24:48], [48:60]
    nav_blocks = [nav_trials[0:24], nav_trials[24:48], nav_trials[48:60]]
    craft_blocks = [craft_trials[0:24], craft_trials[24:48], craft_trials[48:60]]

    if nav_first:
        # S1=nav24, S2=craft24, S3=nav24, S4=craft24, S5=mixed(12n+12c nav-first)
        sessions_spec = [
            (1, "navigation", nav_blocks[0]),
            (2, "crafting", craft_blocks[0]),
            (3, "navigation", nav_blocks[1]),
            (4, "crafting", craft_blocks[1]),
        ]
    else:
        # S1=craft24, S2=nav24, S3=craft24, S4=nav24, S5=mixed(12c+12n craft-first)
        sessions_spec = [
            (1, "crafting", craft_blocks[0]),
            (2, "navigation", nav_blocks[0]),
            (3, "crafting", craft_blocks[1]),
            (4, "navigation", nav_blocks[1]),
        ]

    all_rows: List[Dict[str, object]] = []

    for session_num, domain, block in sessions_spec:
        for idx, trial in enumerate(block, start=1):
            row = dict(trial)
            row["session"] = session_num
            row["session_domain"] = domain
            row["session_index"] = idx
            all_rows.append(row)

    # Session 5 — mixed: interleave nav and craft (12 each = 24 total)
    nav_half = nav_blocks[2]   # 12 trials
    craft_half = craft_blocks[2]  # 12 trials
    mixed: List[Dict[str, object]] = []
    first_half = nav_half if nav_first else craft_half
    second_half = craft_half if nav_first else nav_half
    for i in range(max(len(first_half), len(second_half))):
        if i < len(first_half):
            mixed.append(dict(first_half[i]))
        if i < len(second_half):
            mixed.append(dict(second_half[i]))

    for idx, trial in enumerate(mixed, start=1):
        trial["session"] = 5
        trial["session_domain"] = "mixed"
        trial["session_index"] = idx
        all_rows.append(trial)

    return all_rows


# ── xlsx export ───────────────────────────────────────────────────────────

_COLUMN_ORDER = [
    "session",
    "session_domain",
    "session_index",
    "global_index",
    "block",
    "domain",
    "pair_id",
    "category",
    "start",
    "goal",
    "d_grid",
    "d_loop",
    "grid_paths",
    "loop_paths",
    "multiplicity",
    "block_index",
    "task_id",
    "recipe",
]


def _export_to_xlsx(
    rows: List[Dict[str, object]],
    output_path: Path,
) -> None:
    """Write the schedule rows to an xlsx workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "full_schedule"

    # Determine columns present in the data
    columns = [c for c in _COLUMN_ORDER if any(c in r for r in rows)]

    # Header
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            val = row_data.get(col_name, "")
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-width
    for col_idx, col_name in enumerate(columns, start=1):
        max_len = len(str(col_name))
        for row_idx in range(2, len(rows) + 2):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 3

    # Add a summary sheet
    ws_sum = wb.create_sheet("session_summary")
    sum_headers = ["session", "session_domain", "total_trials", "grid", "loop", "tie"]
    for ci, h in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill

    session_groups: Dict[int, List[Dict[str, object]]] = {}
    for r in rows:
        sn = int(r["session"])
        session_groups.setdefault(sn, []).append(r)

    for ri, sn in enumerate(sorted(session_groups), start=2):
        trials = session_groups[sn]
        cats = {"grid": 0, "loop": 0, "tie": 0}
        for t in trials:
            cats[str(t["category"])] = cats.get(str(t["category"]), 0) + 1
        ws_sum.cell(row=ri, column=1, value=sn)
        ws_sum.cell(row=ri, column=2, value=str(trials[0]["session_domain"]))
        ws_sum.cell(row=ri, column=3, value=len(trials))
        ws_sum.cell(row=ri, column=4, value=cats["grid"])
        ws_sum.cell(row=ri, column=5, value=cats["loop"])
        ws_sum.cell(row=ri, column=6, value=cats["tie"])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


# ── public API ────────────────────────────────────────────────────────────

def generate_and_export(order: Literal["navigation-first", "crafting-first"]) -> Path:
    """Generate the full 120-trial schedule and save as xlsx.

    Returns the path to the generated xlsx file.
    """
    nav_trials = _generate_domain_trials(_NAV_SEED, "navigation")
    craft_trials = _generate_domain_trials(_CRAFT_SEED, "crafting")
    _decorate_crafting_trials(craft_trials, _CRAFT_SEED)

    rows = _build_five_sessions(order, nav_trials, craft_trials)

    filename = f"{order.replace('-', '_')}_schedule.xlsx"
    output_path = _FIXED_SCHEDULES_DIR / filename
    _export_to_xlsx(rows, output_path)
    return output_path


def load_schedule_from_xlsx(xlsx_path: Path | str) -> List[Dict[str, object]]:
    """Read a previously generated schedule xlsx back into a list of trial dicts."""
    from openpyxl import load_workbook

    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb["full_schedule"]

    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h) for h in next(rows_iter)]
    trials: List[Dict[str, object]] = []
    for row_values in rows_iter:
        trial = {}
        for header, value in zip(headers, row_values):
            if value is not None:
                trial[header] = value
        trials.append(trial)
    wb.close()
    return trials


def load_schedule_as_sessions(
    xlsx_path: Path | str,
) -> Dict[str, object]:
    """Load xlsx and restructure into the session-based format expected by
    ``run_experiment_new.py``.

    Returns a dict matching the schema of ``build_session_schedule()`` output.
    """
    trials = load_schedule_from_xlsx(xlsx_path)

    # Group by session
    session_map: Dict[int, List[Dict[str, object]]] = {}
    for trial in trials:
        sn = int(trial["session"])
        session_map.setdefault(sn, []).append(trial)

    sessions: List[Dict[str, object]] = []
    for sn in sorted(session_map):
        session_trials = session_map[sn]
        domain = str(session_trials[0]["session_domain"])

        navigation_trials: List[Dict[str, object]] = []
        crafting_trials: List[Dict[str, object]] = []
        combined_order: List[Dict[str, object]] = []

        for trial in session_trials:
            trial_domain = str(trial["domain"])
            enriched = {
                "index": int(trial.get("session_index", 0)),
                "block_index": int(trial.get("block_index", 0)),
                "pair_id": str(trial.get("pair_id", "")),
                "category": str(trial.get("category", "")),
                "start": int(trial["start"]),
                "goal": int(trial["goal"]),
                "d_grid": int(trial.get("d_grid", 0)),
                "d_loop": int(trial.get("d_loop", 0)),
                "grid_paths": int(trial.get("grid_paths", 0)),
                "loop_paths": int(trial.get("loop_paths", 0)),
                "multiplicity": int(trial.get("multiplicity", 0)),
            }
            if trial_domain == "crafting":
                enriched["task_id"] = str(trial.get("task_id", ""))
                enriched["recipe"] = str(trial.get("recipe", ""))

            if trial_domain == "navigation":
                navigation_trials.append(enriched)
                combined_order.append({"type": "navigation", "trial": enriched})
            else:
                crafting_trials.append(enriched)
                combined_order.append({"type": "crafting", "trial": enriched})

        sessions.append({
            "session": sn,
            "domain": domain,
            "seed": 0,  # N/A — schedule is fully deterministic from xlsx
            "navigation_trials": navigation_trials,
            "crafting_trials": crafting_trials,
            "combined_order": combined_order,
        })

    # Infer order from session 1 domain
    first_domain = str(sessions[0]["domain"]) if sessions else "navigation"
    order = "navigation-first" if first_domain == "navigation" else "crafting-first"

    return {
        "order": order,
        "sessions": sessions,
        "source": str(xlsx_path),
    }


# ── CLI entry point ───────────────────────────────────────────────────────

def main() -> int:
    import argparse

    parser = argparse.ArgumentParser(
        description="Generate fixed trial schedules (xlsx) for the 5-session experiment"
    )
    parser.add_argument(
        "--order",
        choices=["navigation-first", "crafting-first"],
        nargs="*",
        default=["navigation-first", "crafting-first"],
        help="Which order(s) to generate (default: both)",
    )
    args = parser.parse_args()

    for order in args.order:
        path = generate_and_export(order)
        print(f"[OK] {order} -> {path}")

        # Quick verification
        trials = load_schedule_from_xlsx(path)
        sessions_data = load_schedule_as_sessions(path)
        print(f"     Total trials: {len(trials)}")
        for s in sessions_data["sessions"]:
            nav = len(s["navigation_trials"])
            craft = len(s["crafting_trials"])
            cats = {}
            for item in s["combined_order"]:
                cat = str(item["trial"].get("category", ""))
                cats[cat] = cats.get(cat, 0) + 1
            print(
                f"     Session {s['session']} ({s['domain']}): "
                f"nav={nav}, craft={craft}, "
                f"categories={cats}"
            )
        print()

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
